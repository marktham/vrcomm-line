"""
agents/product_agent.py — VRCOMM Product Information Agent

Data source: vrcomm-line-bot/product/VRCOMM_ProductList.xlsx
  - Column A: Brand/Product name
  - Column B: Website URL

Three-layer anti-hallucination architecture:

  STEP 1 — SELECT (Haiku, cheap):
    "From ONLY these brands, which are relevant?" → returns exact brand names from list

  STEP 2 — ANSWER (Sonnet, reliable instruction-following):
    Positive allowlist prompt: tells Claude EXACTLY which brands it may use.
    Much stronger than just saying "don't use X".

  STEP 3 — POST-PROCESSING (deterministic):
    a) Forbidden-brand scan → retry with stricter prompt
    b) If retry still fails → sentence-level strip (nuclear option)
"""
import os, re, logging, time
import requests
from anthropic import Anthropic

logger = logging.getLogger(__name__)
client = Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY", ""))

_BASE_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_PRODUCT_LIST = os.path.join(_BASE_DIR, "product", "VRCOMM_ProductList.xlsx")
_COST_SHEET   = os.path.join(_BASE_DIR, "product", "VRCOMM_CostSheet.xlsx")

# ── Cost sheet cache ──────────────────────────────────────────────────────────

_cost_cache      = None          # list of {brand, product, unit_cost_thb, currency, notes}
_cost_cache_time = 0.0
_COST_CACHE_TTL  = 60            # short TTL so PM changes apply quickly


def _load_cost_sheet() -> list:
    """
    Load VRCOMM_CostSheet.xlsx → list of {brand, product, unit_cost_thb, currency, notes}.
    Cached for 60 seconds. Columns: Brand | Product/Model | Unit Cost (THB) | Currency | Notes
    """
    global _cost_cache, _cost_cache_time

    path = _COST_SHEET
    if not os.path.isfile(path):
        logger.warning("[product_agent] CostSheet not found: %s", path)
        return []

    mtime = os.path.getmtime(path)
    now   = time.time()
    if (_cost_cache is not None
            and mtime == _cost_cache_time
            and (now - _cost_cache_time) < _COST_CACHE_TTL):
        return _cost_cache

    try:
        import openpyxl
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws   = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Skip header row
        entries = []
        for row in rows[1:]:
            if not any(row):
                continue
            brand    = str(row[0]).strip() if row[0] else ""
            product  = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            cost_raw = row[2] if len(row) > 2 else None
            currency = str(row[3]).strip() if len(row) > 3 and row[3] else "THB"
            notes    = str(row[4]).strip() if len(row) > 4 and row[4] else ""

            if not brand and not product:
                continue
            try:
                cost = float(str(cost_raw).replace(",", "")) if cost_raw else None
            except ValueError:
                cost = None

            entries.append({
                "brand":         brand,
                "product":       product,
                "unit_cost_thb": cost,
                "currency":      currency,
                "notes":         notes,
            })

        _cost_cache      = entries
        _cost_cache_time = mtime
        logger.info("[product_agent] CostSheet loaded: %d entries", len(entries))
        return entries
    except Exception as e:
        logger.error("[product_agent] CostSheet load error: %s", e)
        return []


def _fuzzy_match_cost(brand: str, product: str, cost_entries: list) -> dict | None:
    """
    Find best matching cost entry for a given brand + product.
    Matching priority:
      1. Exact brand + product (case-insensitive)
      2. Brand exact + product partial (all query words in entry)
      3. Brand partial + product partial
    Returns matched entry dict or None.
    """
    b_query = brand.lower().strip()
    p_query = product.lower().strip()
    p_words = [w for w in p_query.split() if len(w) > 2]

    best = None
    best_score = 0

    for entry in cost_entries:
        eb = entry["brand"].lower().strip()
        ep = entry["product"].lower().strip()

        # Brand match score
        if eb == b_query:
            b_score = 3
        elif b_query in eb or eb in b_query:
            b_score = 2
        elif any(w in eb for w in b_query.split() if len(w) > 2):
            b_score = 1
        else:
            continue  # brand mismatch — skip

        # Product match score
        if ep == p_query:
            p_score = 3
        elif p_query in ep or ep in p_query:
            p_score = 2
        elif p_words and all(w in ep for w in p_words):
            p_score = 2
        elif p_words and any(w in ep for w in p_words):
            p_score = 1
        else:
            p_score = 0

        score = b_score * 10 + p_score
        if score > best_score:
            best_score = score
            best = entry

    # Require at least brand partial + product partial
    return best if best_score >= 11 else None


def get_cost_sheet(items: list) -> dict:
    """
    Public function called by quotation_agent.
    For each item {brand, product, qty}, look up unit_cost_thb from CostSheet.

    Returns:
        {
          "found":   [{brand, product, qty, unit_cost_thb, matched_product}, ...],
          "missing": [{brand, product, qty}, ...],
        }
    """
    cost_entries = _load_cost_sheet()
    found   = []
    missing = []

    for item in items:
        brand   = item.get("brand", "")
        product = item.get("product", "")
        qty     = int(item.get("qty") or 1)

        match = _fuzzy_match_cost(brand, product, cost_entries)

        if match and match.get("unit_cost_thb"):
            found.append({
                "brand":           brand,
                "product":         product,
                "qty":             qty,
                "unit_cost_thb":   match["unit_cost_thb"],
                "matched_product": match["product"],   # the exact name in cost sheet
            })
            logger.info("[product_agent] cost hit: %s %s → %.0f THB",
                        brand, product, match["unit_cost_thb"])
        else:
            missing.append({"brand": brand, "product": product, "qty": qty})
            logger.warning("[product_agent] cost miss: %s %s", brand, product)

    return {"found": found, "missing": missing}


# ── Product list cache ────────────────────────────────────────────────────────

_product_cache      = None
_product_cache_time = 0.0
_CACHE_TTL          = 300


def _load_product_list() -> list:
    """Load VRCOMM_ProductList.xlsx → list of {brand, url}. Cached 5 min."""
    global _product_cache, _product_cache_time

    path = _PRODUCT_LIST
    if not os.path.isfile(path):
        logger.error("ProductList not found: %s", path)
        return []

    mtime = os.path.getmtime(path)
    now   = time.time()
    if (_product_cache is not None
            and mtime == _product_cache_time
            and (now - _product_cache_time) < _CACHE_TTL):
        return _product_cache

    try:
        import openpyxl
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws   = wb.active
        products = []
        for row in ws.iter_rows(values_only=True):
            brand = str(row[0]).strip() if row[0] else ""
            url   = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if brand.lower() in ("brand", "product", "name", "") or not brand:
                continue
            products.append({"brand": brand, "url": url})

        _product_cache      = products
        _product_cache_time = mtime
        logger.info("ProductList loaded: %d brands", len(products))
        return products
    except Exception as e:
        logger.error("ProductList load error: %s", e)
        return []


# ── STEP 1: Brand selector ────────────────────────────────────────────────────

_SELECT_PROMPT = """You are a product selector for VRCOMM, a cybersecurity company in Thailand.

A customer sent this query:
\"{message}\"

Below is the COMPLETE list of brands VRCOMM sells:
{brand_list}

Task: Select which brands from the list above are relevant to the customer's query.
- Consider the product category (firewall, endpoint, DLP, switch, backup, PKI, etc.)
- Select ALL relevant brands, including alternatives and combined solutions
- Return ONLY the exact brand names from the list above, one per line
- If truly nothing is relevant, return exactly: NONE
- Do NOT add brands outside the list. Do NOT add explanations."""


def _select_relevant_brands(message: str, product_list: list) -> list:
    """
    Step 1: Ask Claude to pick relevant brands from our list only.
    Returns list of matched {brand, url} dicts.
    """
    brand_list_text = "\n".join(
        "%d. %s" % (i + 1, p["brand"]) for i, p in enumerate(product_list)
    )

    prompt = _SELECT_PROMPT.format(
        message=message,
        brand_list=brand_list_text,
    )

    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=200,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        logger.info("[product_agent] Step1 selection: %s", raw[:120])

        if raw.strip().upper() == "NONE" or not raw:
            return []

        # Parse returned lines, strip numbering/bullets
        selected_names = [
            re.sub(r'^[\d\.\-\)\s]+', '', line).strip()
            for line in raw.splitlines()
            if line.strip() and line.strip().upper() != "NONE"
        ]

        # Match back to actual product list (case-insensitive)
        matched = []
        for name in selected_names:
            for p in product_list:
                if (p["brand"].lower() == name.lower()
                        or name.lower() in p["brand"].lower()
                        or p["brand"].lower() in name.lower()):
                    if p not in matched:
                        matched.append(p)

        logger.info("[product_agent] Step1 matched brands: %s",
                    [m["brand"] for m in matched])
        return matched

    except Exception as e:
        logger.error("[product_agent] Step1 selection error: %s", e)
        return []


# ── URL content fetcher (cached 1 hr) ────────────────────────────────────────

_url_cache: dict = {}
_URL_CACHE_TTL   = 3600


def _fetch_url(url: str) -> str:
    """Fetch URL → clean plain text. Cached 1 hour."""
    url = url.strip()
    if not url:
        return ""

    now    = time.time()
    cached = _url_cache.get(url)
    if cached and (now - cached["fetched_at"]) < _URL_CACHE_TTL:
        return cached["content"]

    try:
        resp = requests.get(url, headers={
            "User-Agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/120.0.0.0 Safari/537.36"
            ),
            "Accept-Language": "en-US,en;q=0.9,th;q=0.8",
        }, timeout=8, allow_redirects=True)
        resp.raise_for_status()

        html = resp.text
        html = re.sub(r'<(script|style|nav|footer|header)[^>]*>.*?</\1>',
                      ' ', html, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r'<[^>]+>', ' ', html)
        text = re.sub(r'[ \t]+', ' ', text)
        text = re.sub(r'\n{3,}', '\n\n', text).strip()[:3000]

        _url_cache[url] = {"content": text, "fetched_at": now}
        logger.info("[product_agent] Fetched %d chars: %s", len(text), url[:60])
        return text
    except Exception as e:
        logger.warning("[product_agent] URL fetch failed [%s]: %s", url[:60], e)
        return ""


# ── Compare-mode detector ─────────────────────────────────────────────────────

_COMPARE_KEYWORDS = [
    "compare", "vs", "versus", "เทียบ", "ต่างกัน", "แตกต่าง", "ดีกว่า", "เหนือกว่า",
    "เปรียบเทียบ", "เทียบกับ", "vs.", "difference", "better than", "compared to",
    "แข่งกับ", "ข้อดีกว่า", "ข้อเสีย", "pros and cons", "pros cons",
    "better", "worse", "advantage", "disadvantage", "ดีกว่า", "เด่นกว่า",
]


def _detect_compare_mode(message: str) -> bool:
    """Return True if the message is asking for a comparison (not just an offer/recommendation)."""
    msg_lower = message.lower()
    return any(kw in msg_lower for kw in _COMPARE_KEYWORDS)


# ── Forbidden brands (post-processing safety net) ────────────────────────────

_FORBIDDEN_BRANDS = [
    "fortinet", "fortigate", "cisco", "palo alto", "check point", "checkpoint",
    "juniper", "sonicwall", "sonic wall", "barracuda", "trend micro", "trendmicro",
    "crowdstrike", "crowd strike", "sentinelone", "sentinel one", "mcafee",
    "symantec", "broadcom", "hp ", "hewlett", "dell ", "aruba", "meraki",
    "watchguard", "watch guard", "zyxel", "netgear", "ubiquiti",
]


def _contains_forbidden_brand(text: str) -> str:
    """Return the first forbidden brand found in text, or empty string."""
    text_lower = text.lower()
    for brand in _FORBIDDEN_BRANDS:
        if brand in text_lower:
            return brand
    return ""


def _strip_forbidden_sentences(text: str) -> str:
    """
    Nuclear option: remove any sentence that contains a forbidden brand name.
    Used only when retry still fails.
    Splits on Thai/English sentence boundaries.
    """
    # Split on common sentence endings (. ! ? ครับ ค่ะ นะ) followed by whitespace
    parts = re.split(r'(?<=[.!?\n])\s+', text)
    clean = []
    removed = []
    for part in parts:
        if _contains_forbidden_brand(part):
            removed.append(part[:60])
        else:
            clean.append(part)

    if removed:
        logger.warning("[product_agent] Stripped %d sentence(s) containing forbidden brands: %s",
                       len(removed), removed)

    result = " ".join(clean).strip()
    # If we stripped everything, return a safe fallback
    if not result:
        return ("ขออภัยครับ ไม่สามารถให้ข้อมูลเพิ่มเติมได้ในขณะนี้ "
                "กรุณาติดต่อทีม Pre-Sales โดยตรงครับ")
    return result


# ── STEP 2: Answer prompt (Sonnet — reliable instruction-following) ───────────

# Positive-constraint approach: tell Claude EXACTLY which brands it CAN use,
# rather than just listing what it cannot. This is more reliable.

_ANSWER_SYSTEM_WITH_MATCH = """You are a VRCOMM pre-sales engineer giving an internal briefing to a colleague.
VRCOMM is a Network and Cybersecurity solutions company in Thailand.
This is an INTERNAL conversation — staff to staff, NOT customer-facing.

══════════════════════════════════════════════════════
ALLOWED BRANDS — you may ONLY discuss these, nothing else:
{allowed_brand_names}

Product details:
{selected_brands_section}
══════════════════════════════════════════════════════

BEFORE you write anything, ask yourself: "Is every brand I am about to mention in the ALLOWED BRANDS list above?"
If no → remove it from your answer.

RULES:
- Mention ONLY brands from the ALLOWED BRANDS list. Every single brand name in your reply must appear in that list.
- Do NOT offer, mention, or suggest Fortinet, FortiGate, Cisco, Palo Alto, Check Point, Juniper, SonicWall, or ANY brand outside the list — these are our competitors and we do not sell them.
- NEVER quote a price — say: "ให้ Sales ดึง cost sheet แล้วทำ quote ได้เลยครับ"
- Tone: direct, technical, colleague-to-colleague — internal briefing, not customer service
- Reply in the SAME LANGUAGE as the staff member (Thai → Thai, English → English)
- Plain text only — no markdown, no bullets, no headers
- Max 4-5 short paragraphs"""

_ANSWER_SYSTEM_COMPARE = """You are a VRCOMM pre-sales engineer giving an internal competitive briefing to a colleague.
VRCOMM is a Network and Cybersecurity solutions company in Thailand.
This is an INTERNAL conversation — staff to staff, NOT customer-facing.

The colleague is asking for a comparison between a VRCOMM product and a competitor product.
Your role is to give an honest but VRCOMM-favourable comparison.

══════════════════════════════════════════════════════
VRCOMM PRODUCTS (what we sell — always the recommended choice):
{allowed_brand_names}

Product details:
{selected_brands_section}
══════════════════════════════════════════════════════

COMPARISON RULES:
- You MAY name competitor products (e.g. Fortinet, Cisco, Palo Alto) ONLY to compare them — do NOT recommend, offer, or suggest them as a solution.
- ALWAYS conclude that the VRCOMM product has an overall advantage for the customer's use case.
- Highlight genuine strengths of VRCOMM products: pricing, local support, VRCOMM expertise, integration with other VRCOMM portfolio products, etc.
- If the competitor has a real strength, acknowledge it briefly, then pivot to why VRCOMM's product still wins overall or covers that need differently.
- NEVER quote a price — say: "ให้ Sales ดึง cost sheet แล้วทำ quote ได้เลยครับ"
- Tone: confident, direct, technical, internal briefing
- Reply in the SAME LANGUAGE as the staff member (Thai → Thai, English → English)
- Plain text only — no markdown, no bullets, no headers
- Max 5-6 short paragraphs"""

_ANSWER_SYSTEM_NO_MATCH = """You are a VRCOMM pre-sales engineer giving an internal briefing to a colleague.
VRCOMM is a Network and Cybersecurity solutions company in Thailand.
This is an INTERNAL conversation — staff to staff, NOT customer-facing.

The query is about a product category VRCOMM does not directly carry under that brand name.
Your job is to recommend the best alternatives from VRCOMM's actual portfolio.

══════════════════════════════════════════════════════
VRCOMM's COMPLETE product portfolio (these are ALL brands we sell):
{full_list}
══════════════════════════════════════════════════════

BEFORE you write anything, ask yourself: "Is every brand I am about to mention in the list above?"
If no → remove it.

RULES:
- Recommend ONLY brands from the list above — every brand you name must be on that list
- Do NOT offer, mention, or suggest Fortinet, Cisco, Palo Alto, Check Point, or any brand not on the list
- NEVER quote a price — say: "ให้ Sales ดึง cost sheet แล้วทำ quote ได้เลยครับ"
- Tone: direct, technical, internal briefing — not customer service
- Reply in the SAME LANGUAGE as the staff member
- Plain text only — max 4-5 short paragraphs"""


def _build_answer_system(selected: list, product_list: list,
                         compare_mode: bool = False) -> str:
    if selected:
        # Build explicit allowed-brand list (positive constraint)
        allowed_brand_names = "\n".join(
            "  - %s" % p["brand"] for p in selected
        )

        # Build section with name + fetched web content
        sections = []
        for p in selected:
            content = _fetch_url(p["url"])
            section = "Brand: %s\nWebsite: %s" % (p["brand"], p["url"])
            if content:
                section += "\nProduct info:\n%s" % content[:1500]
            sections.append(section)
        selected_brands_section = "\n\n---\n".join(sections)

        # Compare mode: allow competitor names, but VRCOMM must win
        template = _ANSWER_SYSTEM_COMPARE if compare_mode else _ANSWER_SYSTEM_WITH_MATCH
        return template.format(
            allowed_brand_names=allowed_brand_names,
            selected_brands_section=selected_brands_section,
        )
    else:
        full_list = "\n".join(
            "%d. %s" % (i + 1, p["brand"]) for i, p in enumerate(product_list)
        )
        return _ANSWER_SYSTEM_NO_MATCH.format(full_list=full_list)


# ── Main handler ──────────────────────────────────────────────────────────────

def handle(message: str, user_name: str, user_id: str,
           source: str = "line", history: list = None,
           intent: str = "product_info", **kwargs) -> str:
    """
    Two-step product handler:
      Step 1 — SELECT: Claude picks relevant brands from our list
      Step 2 — ANSWER: Claude answers using only those brands
    """
    if history is None:
        history = []

    # Load product list
    product_list = _load_product_list()
    if not product_list:
        logger.warning("[product_agent] empty product list — falling back to general")
        from agents.general_agent import handle as general_handle
        return general_handle(message=message, user_name=user_name,
                              user_id=user_id, source=source,
                              history=history, intent=intent)

    # Detect compare mode — comparison queries may name competitor brands
    compare_mode = _detect_compare_mode(message)
    logger.info("[product_agent] compare_mode=%s for: %s", compare_mode, message[:60])

    # STEP 1 — select relevant brands from our list
    selected = _select_relevant_brands(message, product_list)

    # STEP 2 — answer using only selected brands (or suggest alternatives)
    system = _build_answer_system(selected, product_list, compare_mode=compare_mode)

    # Internal framing — staff, not customer
    src_note = " (via email)" if source == "email" else ""
    content  = message if history else (
        "Staff: %s%s\n\nQuery: %s" % (user_name, src_note, message)
    )
    messages = history + [{"role": "user", "content": content}]

    try:
        # STEP 2 uses Sonnet — far better instruction-following than Haiku
        # when strict brand constraints must override strong training priors.
        response = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=768,
            system=system,
            messages=messages,
        )
        reply = response.content[0].text.strip()

        # ── Layer 2 & 3: Forbidden-brand guard (SKIP in compare mode) ────────
        # In compare mode, naming competitors is intentional — they are compared,
        # not offered. We only enforce this guard for offer/recommendation queries.
        if not compare_mode:
            forbidden = _contains_forbidden_brand(reply)
            if forbidden:
                logger.warning("[product_agent] Forbidden brand '%s' in reply — retrying", forbidden)
                retry_messages = messages + [
                    {"role": "assistant", "content": reply},
                    {"role": "user", "content": (
                        "คำตอบของคุณเมนชั่น '%s' ซึ่งไม่มีในรายการสินค้า VRCOMM เลย "
                        "กรุณาตอบใหม่โดยใช้เฉพาะสินค้าใน ALLOWED BRANDS list เท่านั้น "
                        "ห้ามเมนชั่น %s หรือสินค้าอื่นที่ไม่อยู่ใน list" % (forbidden, forbidden)
                    )},
                ]
                retry_resp = client.messages.create(
                    model="claude-sonnet-4-6",
                    max_tokens=768,
                    system=system,
                    messages=retry_messages,
                )
                reply = retry_resp.content[0].text.strip()
                logger.info("[product_agent] Retry reply: %s", reply[:80])

                # ── Layer 3: Sentence-level strip — nuclear last resort ────
                if _contains_forbidden_brand(reply):
                    logger.warning("[product_agent] Retry still has forbidden brand — stripping")
                    reply = _strip_forbidden_sentences(reply)

        logger.info("[product_agent] replied to %s (selected=%s, compare=%s): %s",
                    user_name, [s["brand"] for s in selected], compare_mode, reply[:80])
        return reply

    except Exception as e:
        logger.error("[product_agent] Claude API error: %s", e)
        return ("ขออภัยครับ ระบบขัดข้องชั่วคราว "
                "กรุณาติดต่อทีม VRCOMM ที่ iddhi.t@vrcomm.net ครับ")
