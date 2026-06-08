"""
quotation_generator.py — VRCOMM Excel Quotation Generator

Produces a formatted .xlsx quotation file:
  - VRCOMM branded header (navy/gold colour scheme)
  - Quote number: QT-YYYYMMDD-NNN  (auto-incremented per day)
  - Customer info + date + validity
  - Line items table: No | Brand | Product/Model | Qty | Unit Price (THB) | Amount (THB)
  - Subtotal → VAT 7% → Grand Total
  - Footer: prepared by + validity note

Output saved to:  <project_root>/quotations/QT-YYYYMMDD-NNN.xlsx
"""
import os, logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

_BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
_QUOTATIONS_DIR = os.path.join(_BASE_DIR, "quotations")
_COUNTER_FILE   = os.path.join(_QUOTATIONS_DIR, ".quote_counter")

# ── Colour scheme ─────────────────────────────────────────────────────────────
_NAVY   = "1B2A4A"   # header background
_GOLD   = "D4AF37"   # accent / column headers
_WHITE  = "FFFFFF"
_LIGHT  = "EEF2F7"   # alternating row fill
_BORDER = "AAAAAA"

# ── Company info ──────────────────────────────────────────────────────────────
_COMPANY_NAME    = "VRCOMM CO., LTD."
_COMPANY_ADDRESS = "Bangkok, Thailand"
_COMPANY_TEL     = "Tel: +66-2-XXX-XXXX"
_COMPANY_EMAIL   = "sales@vrcomm.net"
_COMPANY_WEBSITE = "www.vrcomm.net"
_TAX_ID          = "Tax ID: 0-1055-56012-43-1"


# ── Quote number generator ────────────────────────────────────────────────────

def _next_quote_number() -> str:
    """
    Generate the next sequential quote number for today: QT-YYYYMMDD-NNN.
    Uses a plain text counter file so it persists across app restarts.
    Thread-safe enough for a single-worker deployment.
    """
    os.makedirs(_QUOTATIONS_DIR, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")

    seq = 1
    if os.path.isfile(_COUNTER_FILE):
        try:
            with open(_COUNTER_FILE, "r") as f:
                parts = f.read().strip().split(":")
            if len(parts) == 2 and parts[0] == today:
                seq = int(parts[1]) + 1
        except Exception:
            pass

    with open(_COUNTER_FILE, "w") as f:
        f.write("%s:%d" % (today, seq))

    return "QT-%s-%03d" % (today, seq)


# ── Style helpers ─────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color=_WHITE, name="Calibri") -> Font:
    return Font(bold=bold, size=size, color=color, name=name)


def _border(style="thin") -> Border:
    s = Side(style=style, color=_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _center(wrap=False) -> Alignment:
    return Alignment(horizontal="center", vertical="center",
                     wrap_text=wrap)


def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


def _set_cell(ws, row, col, value,
              bold=False, size=11, fg=None, font_color="000000",
              align=None, border=False, num_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, color=font_color, name="Calibri")
    if fg:
        cell.fill = _fill(fg)
    if align:
        cell.alignment = align
    if border:
        cell.border = _border()
    if num_format:
        cell.number_format = num_format
    return cell


# ── Main generator ────────────────────────────────────────────────────────────

def generate_quotation(
    customer_name: str,
    items: list,
    margin_pct: float = 30.0,
    validity_days: int = 30,
    notes: str = "",
    prepared_by: str = "",
) -> dict:
    """
    Generate a formatted Excel quotation.

    Args:
        customer_name : customer company name
        items         : list of dicts with keys:
                          brand, product, qty, unit_cost_thb
        margin_pct    : gross margin %. Sell price = cost / (1 - margin/100)
        validity_days : days until quotation expires
        notes         : optional notes printed at bottom
        prepared_by   : staff name

    Returns:
        dict with: quote_no, filepath, grand_total
    """
    os.makedirs(_QUOTATIONS_DIR, exist_ok=True)

    quote_no   = _next_quote_number()
    today      = datetime.now()
    valid_until = today + timedelta(days=validity_days)
    filename   = "%s.xlsx" % quote_no
    filepath   = os.path.join(_QUOTATIONS_DIR, filename)

    # ── Compute pricing ───────────────────────────────────────────────────────
    computed_items = []
    for it in items:
        cost      = float(it.get("unit_cost_thb") or 0)
        qty       = int(it.get("qty") or 1)
        # Sell price = cost / (1 - margin_pct/100)
        if margin_pct >= 100:
            margin_pct = 30.0  # safety
        sell_unit = cost / (1.0 - margin_pct / 100.0)
        amount    = sell_unit * qty
        computed_items.append({
            "brand":      it.get("brand", ""),
            "product":    it.get("product", ""),
            "qty":        qty,
            "unit_cost":  cost,
            "unit_price": round(sell_unit, 2),
            "amount":     round(amount, 2),
        })

    subtotal    = sum(it["amount"] for it in computed_items)
    vat         = round(subtotal * 0.07, 2)
    grand_total = round(subtotal + vat, 2)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # Column widths (A-F)
    col_widths = [6, 20, 40, 10, 18, 18]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── HEADER BLOCK (rows 1-6) ───────────────────────────────────────────────
    ws.row_dimensions[row].height = 14
    ws.merge_cells("A%d:F%d" % (row, row))
    _set_cell(ws, row, 1, "", fg=_NAVY)
    row += 1

    # Company name (large)
    ws.row_dimensions[row].height = 28
    ws.merge_cells("A%d:D%d" % (row, row))
    cell = ws.cell(row=row, column=1, value=_COMPANY_NAME)
    cell.font      = Font(bold=True, size=18, color=_GOLD, name="Calibri")
    cell.fill      = _fill(_NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center",
                               indent=2)

    # Quote label (right side)
    ws.merge_cells("E%d:F%d" % (row, row))
    cell2 = ws.cell(row=row, column=5, value="QUOTATION / ใบเสนอราคา")
    cell2.font      = Font(bold=True, size=12, color=_WHITE, name="Calibri")
    cell2.fill      = _fill(_NAVY)
    cell2.alignment = Alignment(horizontal="right", vertical="center",
                                indent=1)
    row += 1

    # Company address line
    ws.row_dimensions[row].height = 16
    ws.merge_cells("A%d:D%d" % (row, row))
    cell = ws.cell(row=row, column=1,
                   value="%s  |  %s" % (_COMPANY_ADDRESS, _TAX_ID))
    cell.font      = Font(size=9, color="CCCCCC", name="Calibri")
    cell.fill      = _fill(_NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # Quote number (right)
    ws.merge_cells("E%d:F%d" % (row, row))
    cell2 = ws.cell(row=row, column=5, value=quote_no)
    cell2.font      = Font(bold=True, size=11, color=_GOLD, name="Calibri")
    cell2.fill      = _fill(_NAVY)
    cell2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    row += 1

    # Contact line
    ws.row_dimensions[row].height = 14
    ws.merge_cells("A%d:D%d" % (row, row))
    cell = ws.cell(row=row, column=1,
                   value="%s  |  %s  |  %s" % (
                       _COMPANY_TEL, _COMPANY_EMAIL, _COMPANY_WEBSITE))
    cell.font      = Font(size=9, color="CCCCCC", name="Calibri")
    cell.fill      = _fill(_NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)

    # Date (right)
    ws.merge_cells("E%d:F%d" % (row, row))
    cell2 = ws.cell(row=row, column=5,
                    value="Date: %s" % today.strftime("%d %b %Y"))
    cell2.font      = Font(size=9, color=_WHITE, name="Calibri")
    cell2.fill      = _fill(_NAVY)
    cell2.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    row += 1

    # Bottom navy spacer
    ws.row_dimensions[row].height = 10
    ws.merge_cells("A%d:F%d" % (row, row))
    _set_cell(ws, row, 1, "", fg=_NAVY)
    row += 1

    # ── CUSTOMER / DATE INFO (rows 7-10) ──────────────────────────────────────
    row += 1  # blank gap
    ws.row_dimensions[row].height = 18

    label_font = Font(bold=True, size=10, color="1B2A4A", name="Calibri")
    value_font = Font(size=10,            color="222222", name="Calibri")

    def _info_row(r, label, value):
        ws.row_dimensions[r].height = 18
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = label_font
        ws.merge_cells(start_row=r, start_column=2,
                       end_row=r,   end_column=4)
        c2 = ws.cell(row=r, column=2, value=value)
        c2.font = value_font

    _info_row(row,     "Bill To / ลูกค้า :", customer_name)
    _info_row(row + 1, "Prepared By :",       prepared_by or "VRCOMM Sales")
    _info_row(row + 2, "Valid Until :",
              valid_until.strftime("%d %b %Y")
              + " (%d days)" % validity_days)
    row += 4  # skip info rows + 1 gap

    # ── TABLE HEADER ──────────────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    headers = ["No.", "Brand", "Product / Model", "Qty",
               "Unit Price (THB)", "Amount (THB)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font      = Font(bold=True, size=10, color=_WHITE, name="Calibri")
        cell.fill      = _fill(_NAVY)
        cell.alignment = _center()
        cell.border    = _border()
    row += 1

    # ── LINE ITEMS ────────────────────────────────────────────────────────────
    num_fmt_thb  = '#,##0.00'
    for idx, it in enumerate(computed_items, start=1):
        ws.row_dimensions[row].height = 20
        row_fill = _LIGHT if idx % 2 == 0 else _WHITE

        def _td(col, val, align_obj=None, nf=None):
            c = ws.cell(row=row, column=col, value=val)
            c.font      = Font(size=10, color="222222", name="Calibri")
            c.fill      = _fill(row_fill)
            c.border    = _border()
            c.alignment = align_obj or _center()
            if nf:
                c.number_format = nf

        _td(1, idx,                _center())
        _td(2, it["brand"],        _center(wrap=True))
        _td(3, it["product"],      Alignment(horizontal="left",
                                             vertical="center", wrap_text=True))
        _td(4, it["qty"],          _center())
        _td(5, it["unit_price"],   _right(), num_fmt_thb)
        _td(6, it["amount"],       _right(), num_fmt_thb)
        row += 1

    # ── TOTALS BLOCK ──────────────────────────────────────────────────────────
    row += 1  # spacer

    def _total_row(r, label, value, bold=False, bg=None):
        ws.row_dimensions[r].height = 20
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r,   end_column=5)
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font      = Font(bold=bold, size=10,
                            color=(_WHITE if bg == _NAVY else "1B2A4A"),
                            name="Calibri")
        c1.alignment = Alignment(horizontal="right", vertical="center",
                                 indent=1)
        if bg:
            c1.fill = _fill(bg)

        c2 = ws.cell(row=r, column=6, value=value)
        c2.font         = Font(bold=bold, size=10,
                               color=(_WHITE if bg == _NAVY else "222222"),
                               name="Calibri")
        c2.alignment    = _right()
        c2.number_format = num_fmt_thb
        c2.border       = _border()
        if bg:
            c2.fill = _fill(bg)

    _total_row(row,     "Subtotal (ก่อน VAT) :",  subtotal)
    _total_row(row + 1, "VAT 7% :",                vat)
    _total_row(row + 2, "GRAND TOTAL (THB) :",     grand_total,
               bold=True, bg=_NAVY)
    row += 4

    # ── NOTES / FOOTER ────────────────────────────────────────────────────────
    if notes:
        ws.merge_cells("A%d:F%d" % (row, row))
        c = ws.cell(row=row, column=1,
                    value="หมายเหตุ / Notes: %s" % notes)
        c.font      = Font(size=9, color="555555", italic=True, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        row += 1

    row += 1
    ws.merge_cells("A%d:F%d" % (row, row))
    c = ws.cell(row=row, column=1,
                value="ราคานี้มีผลถึง %s  |  Prices are in Thai Baht (THB) and exclusive of VAT unless noted above." % valid_until.strftime("%d %b %Y"))
    c.font      = Font(size=8, color="888888", italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    ws.merge_cells("A%d:F%d" % (row, row))
    c = ws.cell(row=row, column=1,
                value="VRCOMM CO., LTD.  |  %s  |  %s" % (
                    _COMPANY_EMAIL, _COMPANY_WEBSITE))
    c.font      = Font(size=8, color=_GOLD, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Save ──────────────────────────────────────────────────────────────────
    wb.save(filepath)
    logger.info("[quotation_generator] saved: %s | total=%.2f THB", filepath, grand_total)

    return {
        "quote_no":    quote_no,
        "filepath":    filepath,
        "filename":    filename,
        "grand_total": grand_total,
        "subtotal":    subtotal,
        "vat":         vat,
    }
