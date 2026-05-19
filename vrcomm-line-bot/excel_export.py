"""excel_export.py - Export logged LINE messages to Excel"""
import os, logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from db import get_all_messages

logger = logging.getLogger(__name__)
EXPORT_PATH = os.environ.get("EXPORT_PATH", "/tmp/line_messages_export.xlsx")

COLUMNS = [
    ("No.",          "id",           6),
    ("Timestamp",    "timestamp",    22),
    ("Logged At",    "logged_at",    20),
    ("Display Name", "display_name", 20),
    ("User ID",      "user_id",      25),
    ("Source Type",  "source_type",  12),
    ("Source ID",    "source_id",    25),
    ("Msg Type",     "msg_type",     12),
    ("Message",      "msg_text",     50),
    ("Detail",       "msg_detail",   40),
    ("Message ID",   "message_id",   20),
    ("Reply Token",  "reply_token",  40),
]


def export_to_excel(output_path=EXPORT_PATH):
    messages = get_all_messages()
    wb = Workbook()
    ws = wb.active
    ws.title = "LINE Messages"

    ws.merge_cells("A1:L1")
    tc = ws["A1"]
    tc.value = "VRCOMM LINE Bot -- Message Log | Exported: %s" % datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tc.font = Font(bold=True, size=13, color="FFFFFF")
    tc.fill = PatternFill("solid", fgColor="154360")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    thin = Side(border_style="thin", color="AED6F1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill("solid", fgColor="1B4F72")
    h_align = Alignment(horizontal="center", vertical="center")

    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[2].height = 20

    alt_fill = PatternFill("solid", fgColor="EBF5FB")
    for i, msg in enumerate(messages):
        row = i + 3
        fill = alt_fill if i % 2 == 0 else None
        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=col_idx, value=msg.get(field, ""))
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        ws.row_dimensions[row].height = 18

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "VRCOMM LINE Bot -- Summary"
    ws2["A1"].font = Font(bold=True, size=13)
    ws2.merge_cells("A1:D1")
    rows = [("Total Messages", len(messages)),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")), ("", "")]
    sc, mtc, uc = {}, {}, {}
    for m in messages:
        sc[m["source_type"]] = sc.get(m["source_type"], 0) + 1
        mtc[m["msg_type"]] = mtc.get(m["msg_type"], 0) + 1
        uc[m["user_id"]] = 1
    rows.append(("Source Type", "Count"))
    for k, v in sorted(sc.items()):
        rows.append(("  " + k, v))
    rows.append(("", ""))
    rows.append(("Message Type", "Count"))
    for k, v in sorted(mtc.items()):
        rows.append(("  " + k, v))
    rows.append(("", ""))
    rows.append(("Unique Users", len(uc)))
    for r, (a, b) in enumerate(rows, start=3):
        ws2.cell(row=r, column=1, value=a)
        ws2.cell(row=r, column=2, value=b)
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 20

    wb.save(output_path)
    logger.info("Excel exported: %s (%d rows)", output_path, len(messages))
    return output_path
